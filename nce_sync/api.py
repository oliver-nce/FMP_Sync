# Copyright (c) 2026, Oliver Reid and contributors
# For license information, please see license.txt

"""
API endpoints for NCE Sync app.
"""

import frappe
from frappe import _


@frappe.whitelist()
def toggle_auto_sync(table_names):
	"""
	Toggle the auto_sync_active field for the specified WP Tables.

	Args:
		table_names: List of WP Tables names (or JSON string)

	Returns:
		str: Summary message of changes made
	"""
	if isinstance(table_names, str):
		import json

		table_names = json.loads(table_names)

	if not table_names:
		frappe.throw(_("No tables specified"))

	enabled_count = 0
	disabled_count = 0

	for table_name in table_names:
		doc = frappe.get_doc("WP Tables", table_name)

		if doc.auto_sync_active:
			doc.auto_sync_active = 0
			disabled_count += 1
		else:
			doc.auto_sync_active = 1
			enabled_count += 1

		doc.save(ignore_permissions=True)

	frappe.db.commit()

	# Build response message
	parts = []
	if enabled_count:
		parts.append(f"{enabled_count} enabled")
	if disabled_count:
		parts.append(f"{disabled_count} disabled")

	return _("Auto sync: {0}").format(", ".join(parts))


@frappe.whitelist()
def get_table_links_grid_data():
	"""
	Return mirrored tables and their Link field relationships for the grid UI.
	Data is derived from actual DocType metas (no stored copy).

	Returns:
		dict: {
			"tables": [{"doctype": "Events", "label": "Events"}, ...],
			"links": {"Events": {"Venues": [{"field": "venue_id", "label": "Venue"}]}}
		}
	"""
	tables = frappe.get_all(
		"WP Tables",
		filters={"mirror_status": "Mirrored", "frappe_doctype": ["!=", ""]},
		fields=["frappe_doctype", "nce_name", "table_name"],
		order_by="frappe_doctype",
	)

	# Build list of mirrored DocTypes with display labels
	doctypes = []
	seen = set()
	for row in tables:
		dt = row.get("frappe_doctype")
		if not dt or dt in seen:
			continue
		seen.add(dt)
		label = row.get("nce_name") or row.get("table_name") or dt
		doctypes.append({"doctype": dt, "label": label})

	# Scan each DocType's meta for Link fields pointing to other mirrored tables
	mirrored_set = {d["doctype"] for d in doctypes}
	links = {}  # source_doctype -> { target_doctype -> [{"field", "label"}, ...] }

	for d in doctypes:
		source = d["doctype"]
		try:
			meta = frappe.get_meta(source)
		except Exception:
			continue

		for df in meta.fields:
			if df.fieldtype != "Link" or not df.options:
				continue
			target = df.options
			if target not in mirrored_set or target == source:
				continue

			if source not in links:
				links[source] = {}
			if target not in links[source]:
				links[source][target] = []
			links[source][target].append({
				"field": df.fieldname,
				"label": df.label or df.fieldname,
				"many_doctype": source,
			})

			# Also record in the reverse direction so the grid cell works both ways
			if target not in links:
				links[target] = {}
			if source not in links[target]:
				links[target][source] = []
			links[target][source].append({
				"field": df.fieldname,
				"label": df.label or df.fieldname,
				"many_doctype": source,
			})

	return {"tables": doctypes, "links": links}


@frappe.whitelist()
def apply_table_link_changes(to_add, to_delete):
	"""
	Apply pending link field changes in batch.

	Args:
		to_add: JSON list of {"many_doctype", "one_doctype", "field_name"}
		to_delete: JSON list of {"many_doctype", "field_name"}

	Returns:
		str: Summary message
	"""
	import json as _json

	additions = _json.loads(to_add) if isinstance(to_add, str) else to_add
	deletions = _json.loads(to_delete) if isinstance(to_delete, str) else to_delete
	msgs = []

	for item in deletions:
		dt = item["many_doctype"]
		fname = item["field_name"]
		frappe.clear_cache(doctype=dt)
		meta = frappe.get_meta(dt)
		existing = meta.get_field(fname) if meta.has_field(fname) else None
		if not existing or existing.fieldtype != "Link":
			msgs.append(_("{0}.{1} is not a Link field, skipped").format(dt, fname))
			continue
		try:
			doc = frappe.get_doc("DocType", dt)
			for f in doc.fields:
				if f.fieldname == fname:
					f.fieldtype = "Data"
					f.options = ""
					break
			doc.save(ignore_permissions=True)
			frappe.db.commit()
			frappe.clear_cache(doctype=dt)
			msgs.append(_("Reverted {0}.{1} from Link to Data").format(dt, fname))
		except Exception as e:
			msgs.append(_("FAILED revert {0}.{1}: {2}").format(dt, fname, str(e)))

	for item in additions:
		dt = item["many_doctype"]
		one_dt = item["one_doctype"]
		fname = item["field_name"]
		frappe.clear_cache(doctype=dt)
		meta = frappe.get_meta(dt)
		existing = meta.get_field(fname) if meta.has_field(fname) else None

		if existing and existing.fieldtype == "Link" and existing.options == one_dt:
			msgs.append(_("{0}.{1} already links to {2}, skipped").format(dt, fname, one_dt))
			continue

		try:
			doc = frappe.get_doc("DocType", dt)
			if existing:
				for f in doc.fields:
					if f.fieldname == fname:
						f.fieldtype = "Link"
						f.options = one_dt
						break
				action = _("Converted {0}.{1} ({2}) to Link → {3}").format(
					dt, fname, existing.fieldtype, one_dt
				)
			else:
				doc.append("fields", {
					"fieldname": fname,
					"fieldtype": "Link",
					"label": fname.replace("_", " ").title(),
					"options": one_dt,
				})
				action = _("Added {0}.{1} → {2}").format(dt, fname, one_dt)
			doc.save(ignore_permissions=True)
			frappe.db.commit()
			frappe.clear_cache(doctype=dt)
			msgs.append(action)
		except Exception as e:
			msgs.append(_("FAILED {0}.{1}: {2}").format(dt, fname, str(e)))

	return "; ".join(msgs) if msgs else _("No changes applied")


@frappe.whitelist()
def export_all_to_excel(doctype):
	"""
	Export every row and every user-defined column of a DocType to an xlsx file.
	Returns the file URL for browser download.
	"""
	import io

	from openpyxl import Workbook
	from openpyxl.utils import get_column_letter

	if not frappe.has_permission(doctype, "read"):
		frappe.throw(_("Not permitted"), frappe.PermissionError)

	meta = frappe.get_meta(doctype)
	skip_types = frozenset({
		"Section Break", "Column Break", "Tab Break",
		"HTML", "Fold", "Heading",
	})
	fields = [
		df for df in meta.fields
		if df.fieldtype not in skip_types
	]
	fieldnames = ["name"] + [df.fieldname for df in fields]
	labels = ["ID"] + [df.label or df.fieldname for df in fields]

	total = frappe.db.count(doctype)
	frappe.publish_realtime(
		"msgprint",
		{"message": f"Exporting {total} records…", "indicator": "blue", "alert": True},
		user=frappe.session.user,
	)

	rows = frappe.get_all(
		doctype,
		fields=fieldnames,
		limit_page_length=0,
		order_by="name asc",
	)

	wb = Workbook()
	ws = wb.active
	ws.title = doctype[:31]

	ws.append(labels)
	for cell in ws[1]:
		cell.font = cell.font.copy(bold=True)

	for row in rows:
		ws.append([row.get(f) for f in fieldnames])

	for idx, _ in enumerate(labels, 1):
		ws.column_dimensions[get_column_letter(idx)].width = 18

	buf = io.BytesIO()
	wb.save(buf)
	buf.seek(0)

	fname = f"{doctype.replace(' ', '_')}.xlsx"
	file_doc = frappe.get_doc({
		"doctype": "File",
		"file_name": fname,
		"content": buf.getvalue(),
		"is_private": 1,
	})
	file_doc.save(ignore_permissions=True)
	frappe.db.commit()

	frappe.publish_realtime(
		"msgprint",
		{"message": "Done — the file is in your downloads folder", "indicator": "green", "alert": True},
		user=frappe.session.user,
	)

	return file_doc.file_url
